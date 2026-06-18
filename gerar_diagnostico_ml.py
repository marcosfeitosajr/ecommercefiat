from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Iterable, Union

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


REPORT_COLUMNS = [
    "ad_id",
    "title",
    "status",
    "variation",
    "sku",
    "quality",
    "experience",
    "visits",
    "orders",
    "buyers",
    "units",
    "revenue_brl_raw",
    "share_raw",
    "conv_sales_raw",
    "conv_buyers_raw",
    "reviews_total",
    "reviews_bad",
    "reviews_good",
]

ACCENTED = {
    "analise": "análise",
    "anuncios": "anúncios",
    "atuacao": "atuação",
    "acao": "ação",
    "acoes": "ações",
    "basica": "básica",
    "catalogo": "catálogo",
    "concentracao": "concentração",
    "conversao": "conversão",
    "criticos": "críticos",
    "dicionario": "dicionário",
    "diagnostico": "diagnóstico",
    "dispersao": "dispersão",
    "evolucao": "evolução",
    "experiencia": "experiência",
    "historica": "histórica",
    "historico": "histórico",
    "logica": "lógica",
    "media": "média",
    "mes": "mês",
    "metrica": "métrica",
    "oleo": "óleo",
    "operacao": "operação",
    "otimizacao": "otimização",
    "participacao": "participação",
    "projecao": "projeção",
    "rapido": "rápido",
    "tecnica": "técnica",
    "teorico": "teórico",
    "titulo": "título",
    "trafego": "tráfego",
    "tracao": "tração",
    "ultima": "última",
    "unicas": "únicas",
}


@dataclass
class StylePack:
    header_fill: PatternFill
    header_font: Font
    border: Border


def txt(key: str) -> str:
    return ACCENTED.get(key, key)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera um Excel de diagnostico e plano de acao a partir dos relatorios semanais do Mercado Livre."
    )
    parser.add_argument(
        "--reports-dir",
        default="Relatório de anúncios",
        help="Pasta com os relatórios semanais exportados do Mercado Livre.",
    )
    parser.add_argument(
        "--projection-file",
        default="Projeção Sell out - Florença 91219.xlsx",
        help="Planilha opcional com a meta de sell out.",
    )
    parser.add_argument(
        "--market-csv",
        default="Sell out Ecommerce Detalhado por PN.csv",
        help="CSV opcional com benchmark Brasil por PN.",
    )
    parser.add_argument(
        "--output",
        default="Diagnostico_e_Plano_de_Acao_Mercado_Livre_Florenca_Curitiba.xlsx",
        help="Arquivo Excel de saída.",
    )
    parser.add_argument(
        "--top-n",
        type=int,
        default=200,
        help="Quantidade padrão de itens nas abas de ranking e oportunidade.",
    )
    return parser.parse_args()


def parse_period(name: str) -> tuple[pd.Timestamp, pd.Timestamp, int]:
    # O separador entre as duas datas pode ser hifen, underscore ou ausente
    # (alguns exports do ML salvam como 2026_03_012026_03_07.xlsx).
    match = re.search(r"(\d{4}_\d{2}_\d{2})[-_]?(\d{4}_\d{2}_\d{2})", name)
    if not match:
        raise ValueError(
            "Nao foi possivel identificar o periodo no nome do arquivo: "
            f"{name}. Mantenha o nome original exportado pelo Mercado Livre "
            "(deve conter as datas no formato AAAA_MM_DD)."
        )
    start = pd.to_datetime(match.group(1).replace("_", "-"))
    end = pd.to_datetime(match.group(2).replace("_", "-"))
    return start, end, int((end - start).days + 1)


def parse_brl(value: object) -> float:
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float, np.number)):
        return float(value)
    raw = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif re.search(r"^\d{1,3}(\.\d{3})+$", raw):
        raw = raw.replace(".", "")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def parse_count(value: object) -> float:
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, np.integer)):
        return float(value)
    if isinstance(value, (float, np.floating)):
        if not np.isclose(value, round(value)):
            frac = str(value).split(".")[-1]
            if len(frac) == 3 and value < 100:
                return float(round(value * 1000))
        return float(value)
    raw = str(value).strip().replace(" ", "")
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif re.search(r"^\d{1,3}(\.\d{3})+$", raw):
        raw = raw.replace(".", "")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def parse_pct(value: object) -> float:
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float, np.number)):
        return float(value)
    raw = str(value).strip().replace("%", "").replace(".", "").replace(",", ".")
    try:
        return float(raw) / 100.0
    except ValueError:
        return 0.0


def normalize_pn(value: object) -> str:
    raw = str(value).upper().strip()
    raw = re.sub(r"[^A-Z0-9]", "", raw)
    if not raw:
        return ""
    if raw.isdigit():
        return raw.lstrip("0") or "0"
    return raw


def pn_candidates(value: object) -> list[str]:
    raw = str(value).upper().strip()
    if not raw or raw == "NAN":
        return []
    raw = re.sub(r"[^A-Z0-9_ /,+;-]", "", raw)
    parts = re.split(r"[ /,+;]+", raw)
    candidates: list[str] = []
    for part in parts:
        if not part:
            continue
        variants = [part]
        if "_" in part:
            variants.append(part.split("_")[0])
        for variant in variants:
            pn = normalize_pn(variant)
            if pn and pn not in candidates:
                candidates.append(pn)
    return candidates


def _read_report_frame(name: str, source: object) -> pd.DataFrame:
    """Le um unico relatorio semanal (aba 0, cabecalho na linha 6) e anexa o periodo
    derivado do nome do arquivo. `source` pode ser um caminho ou um objeto file-like."""
    start, end, days = parse_period(name)
    df = pd.read_excel(source, sheet_name=0, header=5).dropna(how="all").iloc[:, :18]
    df.columns = REPORT_COLUMNS
    df["period_start"] = start
    df["period_end"] = end
    df["days"] = days
    df["file"] = name
    return df


def _finalize_raw(frames: list[pd.DataFrame]) -> pd.DataFrame:
    raw = pd.concat(frames, ignore_index=True)
    for column in ["visits", "orders", "buyers", "units", "reviews_total", "reviews_bad", "reviews_good"]:
        raw[column] = raw[column].map(parse_count)

    raw["revenue"] = raw["revenue_brl_raw"].map(parse_brl)
    raw["share_pct"] = raw["share_raw"].map(parse_pct)
    raw["conv_sales_report"] = raw["conv_sales_raw"].map(parse_pct)
    raw["conv_buyers_report"] = raw["conv_buyers_raw"].map(parse_pct)
    raw["ad_id"] = raw["ad_id"].astype(str).str.replace(".0", "", regex=False)
    raw["sku_str"] = raw["sku"].astype(str).replace("nan", "")
    raw["pn_candidates"] = raw["sku_str"].map(pn_candidates)
    raw["pn_primary"] = raw["pn_candidates"].map(lambda items: items[0] if items else "")
    status_lower = raw["status"].astype(str).str.lower()
    raw["status_lower"] = status_lower
    raw["active"] = status_lower.str.contains("ativa", na=False) & ~status_lower.str.contains("inativa", na=False)
    return raw


def load_reports(reports_dir: Path) -> pd.DataFrame:
    """Modo pasta (CLI): le todos os .xlsx de um diretorio."""
    files = sorted(reports_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Nenhum relatorio encontrado em {reports_dir}")

    frames = [_read_report_frame(path.name, path) for path in files]
    return _finalize_raw(frames)


def load_reports_from_files(files: Iterable[object]) -> pd.DataFrame:
    """Modo upload (web app): le relatorios de objetos file-like que expoem `.name`
    (ex.: UploadedFile do Streamlit) ou tuplas (nome, bytes/file-like)."""
    frames: list[pd.DataFrame] = []
    for item in files:
        if isinstance(item, tuple):
            name, source = item
        else:
            name = getattr(item, "name", "")
            source = item
        if not name:
            raise ValueError("Cada relatorio precisa ter um nome de arquivo para detectar o periodo.")
        # Garante leitura do inicio caso o buffer ja tenha sido consumido.
        if hasattr(source, "seek"):
            try:
                source.seek(0)
            except (OSError, ValueError):
                pass
        if isinstance(source, (bytes, bytearray)):
            source = BytesIO(source)
        frames.append(_read_report_frame(name, source))

    if not frames:
        raise FileNotFoundError("Nenhum relatorio foi enviado.")
    return _finalize_raw(frames)


def build_weekly(raw: pd.DataFrame) -> pd.DataFrame:
    weekly = (
        raw.groupby(["period_start", "period_end", "days"])
        .agg(
            ads=("ad_id", "count"),
            active_ads=("active", "sum"),
            visits=("visits", "sum"),
            orders=("orders", "sum"),
            buyers=("buyers", "sum"),
            units=("units", "sum"),
            revenue=("revenue", "sum"),
            sold_ads=("orders", lambda s: int((s > 0).sum())),
            visited_ads=("visits", lambda s: int((s > 0).sum())),
            zero_visit_ads=("visits", lambda s: int((s == 0).sum())),
        )
        .reset_index()
    )
    weekly["ads_20vis_no_sale"] = (
        raw.assign(flag=(raw["visits"] >= 20) & (raw["orders"] == 0))
        .groupby(["period_start", "period_end"])["flag"]
        .sum()
        .values
    )
    weekly["conv"] = weekly["orders"] / weekly["visits"].replace(0, np.nan)
    weekly["ticket"] = weekly["revenue"] / weekly["orders"].replace(0, np.nan)
    weekly["revenue_day"] = weekly["revenue"] / weekly["days"]
    weekly["visits_day"] = weekly["visits"] / weekly["days"]
    weekly["orders_day"] = weekly["orders"] / weekly["days"]
    weekly["revenue_per_visit"] = weekly["revenue"] / weekly["visits"].replace(0, np.nan)
    weekly["period"] = weekly["period_start"].dt.strftime("%d/%m") + " a " + weekly["period_end"].dt.strftime("%d/%m")
    return weekly


def build_agg(raw: pd.DataFrame) -> pd.DataFrame:
    agg = (
        raw.groupby(["ad_id", "sku_str", "title"], dropna=False)
        .agg(
            weeks=("file", "nunique"),
            first_week=("period_start", "min"),
            last_week=("period_end", "max"),
            visits=("visits", "sum"),
            orders=("orders", "sum"),
            buyers=("buyers", "sum"),
            units=("units", "sum"),
            revenue=("revenue", "sum"),
            reviews_total=("reviews_total", "max"),
            reviews_bad=("reviews_bad", "max"),
            reviews_good=("reviews_good", "max"),
            weeks_sold=("orders", lambda s: int((s > 0).sum())),
            weeks_visited=("visits", lambda s: int((s > 0).sum())),
        )
        .reset_index()
    )
    agg["conv"] = agg["orders"] / agg["visits"].replace(0, np.nan)
    agg["ticket"] = agg["revenue"] / agg["orders"].replace(0, np.nan)
    agg["revenue_per_visit"] = agg["revenue"] / agg["visits"].replace(0, np.nan)

    latest_end = raw["period_end"].max()
    latest = raw[raw["period_end"] == latest_end].copy()
    latest_by_id = (
        latest.sort_values("revenue", ascending=False)
        .drop_duplicates("ad_id")[
            ["ad_id", "status", "quality", "experience", "active", "visits", "orders", "units", "revenue"]
        ]
        .rename(
            columns={
                "visits": "latest_visits",
                "orders": "latest_orders",
                "units": "latest_units",
                "revenue": "latest_revenue",
            }
        )
    )
    agg = agg.merge(latest_by_id, on="ad_id", how="left")
    return agg


def add_classification(agg: pd.DataFrame, raw: pd.DataFrame) -> tuple[pd.DataFrame, float, float]:
    overall_conv = raw["orders"].sum() / raw["visits"].sum()
    overall_ticket = raw["revenue"].sum() / raw["orders"].sum()

    agg["diagnosis"] = np.select(
        [
            (agg["orders"] > 0) & (agg["conv"] >= 0.10),
            (agg["visits"] >= 50) & (agg["orders"] > 0) & (agg["conv"] < overall_conv),
            (agg["visits"] >= 30) & (agg["orders"] == 0),
            (agg["visits"] == 0),
            (agg["orders"] > 0),
        ],
        [
            "Escalar vencedor",
            "Corrigir conversao",
            "Trafego sem venda",
            "Sem demanda/sem visita",
            "Manter e otimizar",
        ],
        default="Monitorar",
    )
    agg["priority_score"] = (
        agg["revenue"].rank(pct=True).fillna(0) * 35
        + agg["visits"].rank(pct=True).fillna(0) * 25
        + (agg["orders"] > 0).astype(int) * 15
        + ((agg["visits"] >= 30) & (agg["orders"] == 0)).astype(int) * 15
        + (agg["status"].astype(str).str.lower().str.contains("inativa")).astype(int) * 10
    ).round(1)
    agg["potential_revenue"] = 0.0

    low_mask = (agg["visits"] >= 50) & (agg["orders"] > 0) & (agg["conv"] < overall_conv)
    agg.loc[low_mask, "potential_revenue"] = (
        (agg.loc[low_mask, "visits"] * overall_conv - agg.loc[low_mask, "orders"]).clip(lower=0)
        * agg.loc[low_mask, "ticket"]
    ).round(0)

    no_sale_mask = (agg["visits"] >= 30) & (agg["orders"] == 0)
    agg.loc[no_sale_mask, "potential_revenue"] = (
        agg.loc[no_sale_mask, "visits"] * overall_conv * overall_ticket
    ).round(0)
    return agg, overall_conv, overall_ticket


def build_model_view(raw: pd.DataFrame) -> pd.DataFrame:
    keywords = [
        "Toro",
        "Renegade",
        "Compass",
        "Strada",
        "Palio",
        "Uno",
        "Mobi",
        "Ducato",
        "Argo",
        "Fastback",
        "Pulse",
        "Cronos",
        "Siena",
        "Doblo",
        "Punto",
        "Linea",
        "Bravo",
        "Fiorino",
    ]
    rows = []
    for keyword in keywords:
        subset = raw[raw["title"].astype(str).str.contains(keyword, case=False, na=False)]
        if subset.empty:
            continue
        orders = subset["orders"].sum()
        visits = subset["visits"].sum()
        revenue = subset["revenue"].sum()
        rows.append(
            {
                "modelo_palavra_chave": keyword,
                "visits": visits,
                "orders": orders,
                "units": subset["units"].sum(),
                "revenue": revenue,
                "conv": orders / visits if visits else np.nan,
                "ticket": revenue / orders if orders else np.nan,
                "revenue_per_visit": revenue / visits if visits else np.nan,
            }
        )
    return pd.DataFrame(rows).sort_values("revenue", ascending=False)


def _resolve_optional(source: object) -> object:
    """Normaliza uma entrada opcional. Retorna None quando ausente; um caminho que nao
    existe vira None; objetos file-like sao retornados como estao (rebobinados)."""
    if source is None:
        return None
    if isinstance(source, Path):
        return source if source.exists() else None
    if hasattr(source, "seek"):
        try:
            source.seek(0)
        except (OSError, ValueError):
            pass
    return source


def load_market_csv(csv_source: object = None) -> pd.DataFrame:
    src = _resolve_optional(csv_source)
    if src is None:
        return pd.DataFrame()

    market = pd.read_csv(src)
    rename_map = {
        "DESENHO": "pn",
        "DESCRIÇÃO": "descricao",
        "TIPOLOGIA": "tipologia",
        "SELLOUT": "sellout_brasil",
        "QTD": "qty_brasil",
        "TICKET MÉDIO": "ticket_brasil_raw",
    }
    market = market.rename(columns=rename_map)
    keep_cols = ["pn", "descricao", "tipologia", "sellout_brasil", "qty_brasil", "ticket_brasil_raw"]
    market = market[keep_cols].copy()
    market["pn_original"] = market["pn"].astype(str)
    market["pn"] = market["pn"].map(normalize_pn)
    market["sellout_brasil"] = pd.to_numeric(market["sellout_brasil"], errors="coerce").fillna(0.0)
    market["qty_brasil"] = pd.to_numeric(market["qty_brasil"], errors="coerce").fillna(0.0)
    market["ticket_brasil"] = market["ticket_brasil_raw"].map(parse_brl)
    market = market.drop(columns=["ticket_brasil_raw"])
    market = market.groupby("pn", as_index=False).agg(
        pn_original=("pn_original", "first"),
        descricao=("descricao", "first"),
        tipologia=("tipologia", "first"),
        sellout_brasil=("sellout_brasil", "sum"),
        qty_brasil=("qty_brasil", "sum"),
        ticket_brasil=("ticket_brasil", "mean"),
    )
    return market


def build_market_benchmark(raw: pd.DataFrame, market: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if market.empty:
        return pd.DataFrame(), pd.DataFrame()

    raw["pn_match_count"] = raw["pn_candidates"].map(len)
    multi_match = raw[raw["pn_match_count"] > 1].copy()
    multi_match_view = (
        multi_match[
            [
                "ad_id",
                "sku_str",
                "pn_primary",
                "title",
                "pn_match_count",
                "visits",
                "orders",
                "units",
                "revenue",
            ]
        ]
        .assign(pn_candidates_text=multi_match["pn_candidates"].map(lambda items: ", ".join(items)))
        .sort_values(["pn_match_count", "revenue"], ascending=[False, False])
    )

    florida = (
        raw[raw["pn_primary"] != ""]
        .groupby("pn_primary", as_index=False)
        .agg(
            florenca_ads=("ad_id", "nunique"),
            florenca_visits=("visits", "sum"),
            florenca_orders=("orders", "sum"),
            florenca_units=("units", "sum"),
            florenca_revenue=("revenue", "sum"),
        )
        .rename(columns={"pn_primary": "pn"})
    )
    florida["ticket_florenca"] = florida["florenca_revenue"] / florida["florenca_orders"].replace(0, np.nan)
    florida["receita_por_unidade_florenca"] = florida["florenca_revenue"] / florida["florenca_units"].replace(0, np.nan)

    benchmark = market.merge(florida, on="pn", how="left")
    fill_zero_cols = ["florenca_ads", "florenca_visits", "florenca_orders", "florenca_units", "florenca_revenue"]
    for column in fill_zero_cols:
        benchmark[column] = benchmark[column].fillna(0.0)

    benchmark["ticket_florenca"] = benchmark["ticket_florenca"].fillna(0.0)
    benchmark["receita_por_unidade_florenca"] = benchmark["receita_por_unidade_florenca"].fillna(0.0)
    benchmark["share_qty_florenca"] = benchmark["florenca_units"] / benchmark["qty_brasil"].replace(0, np.nan)
    benchmark["share_sellout_florenca"] = benchmark["florenca_revenue"] / benchmark["sellout_brasil"].replace(0, np.nan)
    benchmark["delta_ticket_pct"] = benchmark["ticket_florenca"] / benchmark["ticket_brasil"].replace(0, np.nan) - 1
    benchmark["gap_units_brasil_vs_florenca"] = benchmark["qty_brasil"] - benchmark["florenca_units"]
    benchmark["gap_revenue_brasil_vs_florenca"] = benchmark["sellout_brasil"] - benchmark["florenca_revenue"]
    benchmark["benchmark_status"] = np.select(
        [
            (benchmark["florenca_units"] == 0) & (benchmark["qty_brasil"] > 0),
            (benchmark["share_qty_florenca"].fillna(0) < 0.01) & (benchmark["qty_brasil"] >= 50),
            benchmark["delta_ticket_pct"].fillna(0) > 0.15,
            benchmark["delta_ticket_pct"].fillna(0) < -0.15,
        ],
        [
            "Sem venda Florenca",
            "Baixa participacao Florenca",
            "Preco Florenca acima mercado",
            "Preco Florenca abaixo mercado",
        ],
        default="Acompanhar",
    )
    benchmark = benchmark.sort_values(["sellout_brasil", "qty_brasil"], ascending=False)
    return benchmark, multi_match_view


def load_projection(projection_source: object = None) -> tuple[pd.DataFrame, float, float]:
    base_apr = 330000.0
    target_dec = 1500000.0
    src = _resolve_optional(projection_source)
    if src is not None:
        projection_sheet = pd.read_excel(src, sheet_name=0, header=None)
        if pd.notna(projection_sheet.iloc[5, 5]):
            base_apr = float(projection_sheet.iloc[5, 5])
        if pd.notna(projection_sheet.iloc[5, 6]):
            target_dec = float(projection_sheet.iloc[5, 6])

    monthly_rate = math.pow(target_dec / base_apr, 1 / 8) - 1
    rows = []
    for index, month in enumerate(["Abr/26", "Mai/26", "Jun/26", "Jul/26", "Ago/26", "Set/26", "Out/26", "Nov/26", "Dez/26"]):
        monthly_goal = base_apr * math.pow(1 + monthly_rate, index)
        rows.append(
            {
                "mes": month,
                "meta_mensal": monthly_goal,
                "meta_semanal_media": monthly_goal / 4.345,
                "crescimento_mensal_necessario": monthly_rate,
            }
        )
    return pd.DataFrame(rows), base_apr, target_dec


def build_summary(
    raw: pd.DataFrame,
    agg: pd.DataFrame,
    weekly: pd.DataFrame,
    projection_target: float,
    monthly_rate: float,
    benchmark: pd.DataFrame,
) -> pd.DataFrame:
    sold = agg[agg["revenue"] > 0].sort_values("revenue", ascending=False).copy()
    sold["cum_revenue"] = sold["revenue"].cumsum()
    sold["cum_share"] = sold["cum_revenue"] / sold["revenue"].sum()
    pareto_ads_80 = int((sold["cum_share"] <= 0.8).sum() + 1) if not sold.empty else 0

    low_conv = agg[(agg["visits"] >= 50) & (agg["orders"] > 0) & (agg["conv"] < (raw["orders"].sum() / raw["visits"].sum()))]
    no_sale = agg[(agg["visits"] >= 30) & (agg["orders"] == 0)]
    inactive = agg[(agg["status"].astype(str).str.lower().str.contains("inativa")) & ((agg["visits"] >= 20) | (agg["orders"] > 0))]

    rows = [
        ("Periodo analisado", f"{raw['period_start'].min():%d/%m/%Y} a {raw['period_end'].max():%d/%m/%Y}", "Base lida a partir dos relat" + "órios semanais."),
        ("Receita bruta ML", raw["revenue"].sum(), "Soma das vendas brutas dos relat" + "órios."),
        ("Unidades vendidas", raw["units"].sum(), ""),
        ("Pedidos", raw["orders"].sum(), ""),
        ("Compradores unicos", raw["buyers"].sum(), ""),
        ("Visitas unicas", raw["visits"].sum(), ""),
        ("Conversao media", raw["orders"].sum() / raw["visits"].sum(), "Pedidos / visitas."),
        ("Ticket medio", raw["revenue"].sum() / raw["orders"].sum(), "Receita / pedidos."),
        ("Receita por visita", raw["revenue"].sum() / raw["visits"].sum(), "Qualidade economica do trafego."),
        ("Anuncios unicos analisados", raw["ad_id"].nunique(), ""),
        ("Anuncios com venda", int((agg["orders"] > 0).sum()), f"{(agg['orders'] > 0).mean():.1%} da base unica."),
        ("Anuncios com zero visita no periodo", int((agg["visits"] == 0).sum()), ""),
        ("Anuncios ativos na ultima semana", int(raw.loc[raw['period_end'] == raw['period_end'].max(), 'active'].sum()), ""),
        ("Anuncios inativos na ultima semana", int((~raw.loc[raw['period_end'] == raw['period_end'].max(), 'active']).sum()), ""),
        ("Anuncios vendidos na ultima semana", int((raw.loc[raw['period_end'] == raw['period_end'].max(), 'orders'] > 0).sum()), ""),
        ("Top 10 anuncios: participacao receita", sold.head(10)["revenue"].sum() / sold["revenue"].sum(), ""),
        ("Top 20 anuncios: participacao receita", sold.head(20)["revenue"].sum() / sold["revenue"].sum(), ""),
        ("Qtd anuncios para 80% da receita", pareto_ads_80, ""),
        ("Anuncios 30+ visitas sem venda", len(no_sale), f"Potencial teorico: R$ {no_sale['potential_revenue'].sum():,.0f}"),
        ("Anuncios 50+ visitas baixa conversao", len(low_conv), f"Potencial teorico: R$ {low_conv['potential_revenue'].sum():,.0f}"),
        ("Inativos com demanda/venda historica", len(inactive), f"Receita historica: R$ {inactive['revenue'].sum():,.0f}"),
        (
            "PNs benchmark com venda Brasil",
            int((benchmark["qty_brasil"] > 0).sum()) if not benchmark.empty else 0,
            "Comparativo por PN vindo do CSV Brasil." if not benchmark.empty else "CSV Brasil nao encontrado.",
        ),
        (
            "PNs sem venda Florenca mas com venda Brasil",
            int(((benchmark["qty_brasil"] > 0) & (benchmark["florenca_units"] == 0)).sum()) if not benchmark.empty else 0,
            "Prioridade de expansao de sortimento." if not benchmark.empty else "",
        ),
        (
            "PNs com preco Florenca acima do mercado",
            int((benchmark["delta_ticket_pct"].fillna(0) > 0.15).sum()) if not benchmark.empty else 0,
            "Usar junto com a aba de benchmark." if not benchmark.empty else "",
        ),
        ("Meta dezembro indicada", projection_target, "Planilha de projecao."),
        ("Crescimento mensal necessario", monthly_rate, "Abril a dezembro."),
    ]
    return pd.DataFrame(rows, columns=["indicador", "valor", "comentario"])


def build_notes() -> pd.DataFrame:
    rows = [
        ("Diagnostico geral", "O canal tem tracao, mas a base e muito desigual: poucos anuncios concentram a receita e ha muitos itens com trafego improdutivo ou sem visita."),
        ("Queda recente", "A ultima semana do recorte precisa ser lida com cautela porque tem menos dias, mas ainda assim a receita diaria enfraqueceu versus as semanas anteriores."),
        ("Concentracao", "Poucos SKUs sustentam grande parte do faturamento. Esses itens precisam de protecao de estoque, preco e ranking."),
        ("Conversao", "Os maiores ganhos nao estao so em gerar mais visitas; boa parte do potencial esta em corrigir anuncios que ja recebem trafego."),
        ("Inativos", "Ha anuncios pausados com historico de venda. Reativar esses itens costuma ser o ganho mais rapido."),
        ("Benchmark Brasil", "O CSV por PN permite medir participacao Florenca no volume Brasil e comparar ticket medio da Florenca com o ticket medio de mercado."),
        ("Meta", "A meta de dezembro exige crescimento mensal forte. O acompanhamento semanal ajuda a antecipar desvios."),
    ]
    return pd.DataFrame(rows, columns=["tema", "leitura"])


def build_action_plan() -> pd.DataFrame:
    rows = [
        [1, "Recuperar receita parada", "Reativar ou recriar anuncios inativos com demanda historica.", "09_Inativos_Demanda", "7 dias", "Alta", "E-commerce + estoque", "Ruptura, pausa por estoque, preco/frete ou ficha.", "Itens reativados, receita recuperada, vendas por SKU."],
        [2, "Corrigir conversao", "Auditar anuncios com 50+ visitas e conversao abaixo da media.", "07_Gargalos_Conversao", "15 dias", "Alta", "E-commerce + precificacao", "Preco, frete, foto principal, titulo e compatibilidade.", "Conversao por anuncio e receita por visita."],
        [3, "Transformar trafego em venda", "Atacar anuncios com 30+ visitas e zero venda.", "08_Visitas_Sem_Venda", "15 dias", "Alta", "E-commerce + produto", "Anuncio fraco, produto inadequado, preco fora ou falta de prova.", "Primeira venda, conversao e queda de visitas improdutivas."],
        [4, "Escalar vencedores", "Garantir estoque, FULL e kits dos itens com alta conversao.", "10_Escalar_Vencedores", "30 dias", "Alta", "Estoque + marketplace", "Ruptura, perda de ranking e margem sem controle.", "Receita top SKUs, disponibilidade e participacao FULL."],
        [5, "Melhorar qualidade da base", "Elevar anuncios basicos com fotos, ficha tecnica, compatibilidade e atributos.", "12_Base_Consolidada", "30 dias", "Media", "Catalogo + e-commerce", "Qualidade basica limita confianca e performance.", "Qualidade do anuncio, experiencia e visitas."],
        [6, "Sanear cauda longa", "Revisar anuncios sem visita para corrigir, consolidar ou descontinuar.", "12_Base_Consolidada", "45 dias", "Media", "E-commerce", "Base dispersa energia operacional.", "Anuncios zero visita e visitas por anuncio."],
        [7, "Gestao da meta", "Comparar receita semanal com a meta do ano e agir sobre o gap.", "03_Metas", "Semanal", "Alta", "Gestao", "Meta de dezembro exige crescimento mensal acelerado.", "Receita semanal versus meta e gap acumulado."],
    ]
    columns = ["ordem", "frente", "acao", "aba_base", "prazo", "prioridade", "responsavel_sugerido", "diagnostico", "kpi"]
    return pd.DataFrame(rows, columns=columns)


def write_sheet(ws, df: pd.DataFrame, styles: StylePack, title: str, subtitle: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="666666")
    start_row = 4

    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws.cell(start_row, col_idx, column)
        cell.fill = styles.header_fill
        cell.font = styles.header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles.border

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            if isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            if not isinstance(value, str) and pd.isna(value):
                value = None
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = styles.border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            apply_number_format(cell, df.columns[col_idx - 1], value)

    if not df.empty:
        ref = f"A{start_row}:{get_column_letter(len(df.columns))}{start_row + len(df)}"
        table = Table(displayName=clean_table_name(ws.title), ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        ws.auto_filter.ref = ref

    ws.freeze_panes = "A5"
    fit_columns(ws, df)


def apply_number_format(cell, column: str, value: object) -> None:
    money_columns = {
        "revenue",
        "ticket",
        "revenue_day",
        "revenue_per_visit",
        "potential_revenue",
        "latest_revenue",
        "meta_mensal",
        "meta_semanal_media",
        "sellout_brasil",
        "ticket_brasil",
        "florenca_revenue",
        "ticket_florenca",
        "receita_por_unidade_florenca",
        "gap_revenue_brasil_vs_florenca",
    }
    pct_columns = {
        "conv",
        "crescimento_mensal_necessario",
        "valor",
        "share_qty_florenca",
        "share_sellout_florenca",
        "delta_ticket_pct",
    }
    int_columns = {
        "visits",
        "orders",
        "buyers",
        "units",
        "ads",
        "active_ads",
        "sold_ads",
        "visited_ads",
        "zero_visit_ads",
        "ads_20vis_no_sale",
        "latest_visits",
        "latest_orders",
        "latest_units",
        "weeks",
        "weeks_sold",
        "weeks_visited",
        "reviews_total",
        "reviews_bad",
        "reviews_good",
        "priority_score",
        "valor",
        "qty_brasil",
        "florenca_ads",
        "florenca_visits",
        "florenca_orders",
        "florenca_units",
        "gap_units_brasil_vs_florenca",
        "pn_match_count",
    }
    if isinstance(value, datetime):
        cell.number_format = "dd/mm/yyyy"
        return
    if not isinstance(value, (int, float, np.number)):
        return

    if column in money_columns:
        cell.number_format = 'R$ #,##0'
    elif column in pct_columns and abs(float(value)) <= 1.0:
        cell.number_format = "0.0%"
    elif column in int_columns:
        cell.number_format = "#,##0"


def fit_columns(ws, df: pd.DataFrame) -> None:
    wide = {"title", "acao", "diagnostico", "kpi", "comentario", "leitura", "descricao", "pn_candidates_text"}
    medium = {"ad_id", "sku_str", "status", "quality", "experience", "diagnosis", "responsavel_sugerido", "tipologia", "benchmark_status"}
    for idx, column in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        if column in wide:
            ws.column_dimensions[letter].width = 48
        elif column in medium:
            ws.column_dimensions[letter].width = 18
        else:
            ws.column_dimensions[letter].width = min(max(len(str(column)) + 2, 12), 22)


def clean_table_name(title: str) -> str:
    table_name = re.sub(r"[^A-Za-z0-9_]", "", title)
    if not table_name:
        table_name = "Tabela"
    if not table_name[0].isalpha():
        table_name = "T" + table_name
    return table_name[:240]


def add_weekly_charts(ws, row_count: int) -> None:
    max_row = 4 + row_count

    revenue_chart = BarChart()
    revenue_chart.title = "Receita semanal"
    revenue_chart.y_axis.title = "Receita (R$)"
    revenue_chart.x_axis.title = "Semana"
    revenue_data = Reference(ws, min_col=11, min_row=4, max_row=max_row)
    categories = Reference(ws, min_col=1, min_row=5, max_row=max_row)
    revenue_chart.add_data(revenue_data, titles_from_data=True)
    revenue_chart.set_categories(categories)
    revenue_chart.height = 7
    revenue_chart.width = 16
    ws.add_chart(revenue_chart, "W5")

    conv_chart = LineChart()
    conv_chart.title = "Conversao semanal"
    conv_chart.y_axis.title = "Conversao"
    conv_data = Reference(ws, min_col=12, min_row=4, max_row=max_row)
    conv_chart.add_data(conv_data, titles_from_data=True)
    conv_chart.set_categories(categories)
    conv_chart.height = 7
    conv_chart.width = 16
    ws.add_chart(conv_chart, "W20")


def apply_conditional_formatting(ws) -> None:
    headers = [ws.cell(4, idx).value for idx in range(1, ws.max_column + 1)]
    for target in ["revenue", "potential_revenue", "priority_score", "conv"]:
        if target not in headers:
            continue
        col_idx = headers.index(target) + 1
        rng = f"{get_column_letter(col_idx)}5:{get_column_letter(col_idx)}{ws.max_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B" if target == "conv" else "FFFFFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )


def build_workbook(
    output: Union[Path, str, BytesIO],
    raw: pd.DataFrame,
    weekly: pd.DataFrame,
    agg: pd.DataFrame,
    models: pd.DataFrame,
    projection: pd.DataFrame,
    summary: pd.DataFrame,
    notes: pd.DataFrame,
    action_plan: pd.DataFrame,
    benchmark: pd.DataFrame,
    multi_match: pd.DataFrame,
    top_n: int,
) -> None:
    sold = agg[agg["revenue"] > 0].sort_values("revenue", ascending=False).copy()
    low_conv = agg[(agg["visits"] >= 50) & (agg["orders"] > 0) & (agg["conv"] < (raw["orders"].sum() / raw["visits"].sum()))]
    no_sale = agg[(agg["visits"] >= 30) & (agg["orders"] == 0)]
    inactive = agg[(agg["status"].astype(str).str.lower().str.contains("inativa")) & ((agg["visits"] >= 20) | (agg["orders"] > 0))]
    winners = agg[(agg["visits"] >= 10) & (agg["conv"] >= 0.10) & (agg["orders"] >= 2)]
    base = agg.sort_values("priority_score", ascending=False)

    sheets = {
        "00_Resumo": summary,
        "01_Diagnostico": notes,
        "02_Semanal": weekly[
            [
                "period",
                "period_start",
                "period_end",
                "days",
                "ads",
                "active_ads",
                "visits",
                "orders",
                "buyers",
                "units",
                "revenue",
                "conv",
                "ticket",
                "revenue_day",
                "visits_day",
                "orders_day",
                "revenue_per_visit",
                "sold_ads",
                "visited_ads",
                "zero_visit_ads",
                "ads_20vis_no_sale",
            ]
        ],
        "03_Metas": projection,
        "04_Plano_Acao": action_plan,
        "05_Top_Receita": sold.head(top_n)[
            [
                "ad_id",
                "sku_str",
                "title",
                "status",
                "quality",
                "experience",
                "visits",
                "orders",
                "buyers",
                "units",
                "revenue",
                "conv",
                "ticket",
                "revenue_per_visit",
                "weeks_sold",
                "weeks_visited",
                "latest_visits",
                "latest_orders",
                "latest_revenue",
                "priority_score",
                "diagnosis",
            ]
        ],
        "06_Top_Unidades": agg.sort_values(["units", "revenue"], ascending=False).head(top_n)[
            [
                "ad_id",
                "sku_str",
                "title",
                "status",
                "quality",
                "experience",
                "visits",
                "orders",
                "units",
                "revenue",
                "conv",
                "ticket",
                "weeks_sold",
                "diagnosis",
            ]
        ],
        "07_Gargalos_Conversao": low_conv.sort_values("potential_revenue", ascending=False).head(top_n)[
            [
                "ad_id",
                "sku_str",
                "title",
                "status",
                "quality",
                "experience",
                "visits",
                "orders",
                "units",
                "revenue",
                "conv",
                "ticket",
                "potential_revenue",
                "latest_visits",
                "latest_orders",
                "priority_score",
            ]
        ],
        "08_Visitas_Sem_Venda": no_sale.sort_values(["visits", "potential_revenue"], ascending=False).head(top_n)[
            [
                "ad_id",
                "sku_str",
                "title",
                "status",
                "quality",
                "experience",
                "visits",
                "orders",
                "units",
                "revenue",
                "potential_revenue",
                "latest_visits",
                "reviews_total",
                "priority_score",
            ]
        ],
        "09_Inativos_Demanda": inactive.sort_values(["revenue", "visits"], ascending=False).head(top_n)[
            [
                "ad_id",
                "sku_str",
                "title",
                "status",
                "quality",
                "experience",
                "visits",
                "orders",
                "units",
                "revenue",
                "conv",
                "ticket",
                "latest_visits",
                "latest_orders",
                "priority_score",
            ]
        ],
        "10_Escalar_Vencedores": winners.sort_values(["revenue", "conv"], ascending=False).head(top_n)[
            [
                "ad_id",
                "sku_str",
                "title",
                "status",
                "quality",
                "experience",
                "visits",
                "orders",
                "units",
                "revenue",
                "conv",
                "ticket",
                "revenue_per_visit",
                "weeks_sold",
                "latest_visits",
                "latest_orders",
                "priority_score",
            ]
        ],
        "11_Modelos": models,
        "12_Base_Consolidada": base[
            [
                "ad_id",
                "sku_str",
                "title",
                "status",
                "quality",
                "experience",
                "weeks",
                "visits",
                "orders",
                "buyers",
                "units",
                "revenue",
                "conv",
                "ticket",
                "revenue_per_visit",
                "reviews_total",
                "reviews_bad",
                "reviews_good",
                "weeks_sold",
                "weeks_visited",
                "latest_visits",
                "latest_orders",
                "latest_revenue",
                "diagnosis",
                "potential_revenue",
                "priority_score",
            ]
        ],
    }
    if not benchmark.empty:
        sheets["13_Benchmark_PN"] = benchmark[
            [
                "pn_original",
                "descricao",
                "tipologia",
                "qty_brasil",
                "sellout_brasil",
                "ticket_brasil",
                "florenca_ads",
                "florenca_visits",
                "florenca_orders",
                "florenca_units",
                "florenca_revenue",
                "ticket_florenca",
                "share_qty_florenca",
                "share_sellout_florenca",
                "delta_ticket_pct",
                "gap_units_brasil_vs_florenca",
                "gap_revenue_brasil_vs_florenca",
                "benchmark_status",
            ]
        ]
        sheets["14_PN_MultiMatch"] = multi_match[
            [
                "ad_id",
                "sku_str",
                "pn_primary",
                "pn_candidates_text",
                "title",
                "pn_match_count",
                "visits",
                "orders",
                "units",
                "revenue",
            ]
        ]

    wb = Workbook()
    wb.remove(wb.active)
    styles = StylePack(
        header_fill=PatternFill("solid", fgColor="1F4E78"),
        header_font=Font(color="FFFFFF", bold=True),
        border=Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        ),
    )

    subtitle = f"Gerado em {datetime.now():%d/%m/%Y %H:%M} | Mercado Livre | Evolucao anual"
    for title, df in sheets.items():
        ws = wb.create_sheet(title)
        write_sheet(ws, df, styles, title.replace("_", " "), subtitle)
        if title == "02_Semanal":
            add_weekly_charts(ws, len(df))
        if title in {
            "05_Top_Receita",
            "07_Gargalos_Conversao",
            "08_Visitas_Sem_Venda",
            "09_Inativos_Demanda",
            "10_Escalar_Vencedores",
            "12_Base_Consolidada",
            "13_Benchmark_PN",
        }:
            apply_conditional_formatting(ws)
        ws.sheet_view.showGridLines = False

    add_quick_read(wb["00_Resumo"], styles)
    add_dictionary_sheet(wb, styles)
    wb.save(output)


def add_quick_read(ws, styles: StylePack) -> None:
    ws["E4"] = "Leitura rapida"
    ws["E4"].fill = styles.header_fill
    ws["E4"].font = styles.header_font
    notes = [
        "1) Proteger os top SKUs de oleo, aditivo e kits que concentram a receita.",
        "2) Reativar anuncios inativos com demanda historica antes de mexer em cauda longa.",
        "3) Corrigir anuncios com muitas visitas e baixa conversao: preco, frete, foto, titulo e compatibilidade.",
        "4) Usar o benchmark por PN para achar onde a Florenca vende pouco versus o mercado Brasil ou pratica ticket desalinhado.",
        "5) Sanear a base sem visita para reduzir dispersao operacional.",
        "6) Acompanhar a meta semanal para antecipar desvios do plano anual.",
    ]
    for idx, note in enumerate(notes, start=5):
        cell = ws.cell(idx, 5, note)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = styles.border
    ws.column_dimensions["E"].width = 72


def add_dictionary_sheet(wb: Workbook, styles: StylePack) -> None:
    sheet_name = "15_Dicionario" if "13_Benchmark_PN" in wb.sheetnames else "13_Dicionario"
    ws = wb.create_sheet(sheet_name)
    rows = [
        ("Conversao", "Pedidos / visitas unicas."),
        ("Ticket", "Receita bruta / pedidos."),
        ("Receita por visita", "Receita bruta / visitas unicas."),
        ("Potencial teorico", "Anuncios com baixa conversao sao simulados na conversao media; anuncios sem venda usam conversao media e ticket medio geral."),
        ("Escalar vencedor", "Anuncio com pelo menos 10 visitas, 2 pedidos e conversao igual ou superior a 10%."),
        ("Corrigir conversao", "Anuncio com 50 ou mais visitas, venda realizada, mas conversao abaixo da media geral."),
        ("Trafego sem venda", "Anuncio com 30 ou mais visitas e zero pedido no periodo."),
        ("Inativos com demanda", "Anuncios inativos na ultima semana, mas com visita ou venda relevante no historico."),
        ("Benchmark PN", "Comparacao por PN entre volume Brasil do CSV e volume Florenca calculado pelos anuncios."),
        ("MultiMatch", "SKUs que apontam para mais de um PN possivel. Use essa aba para auditoria manual dos casos ambiguos."),
        ("Leitura do ano", "A atualizacao do arquivo depende apenas da entrada de novos relatorios semanais na pasta informada."),
    ]
    ws["A1"] = "Dicionario e criterios"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws["A2"] = f"Gerado em {datetime.now():%d/%m/%Y %H:%M}"
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A4"] = "tema"
    ws["B4"] = "leitura"
    for coord in ["A4", "B4"]:
        ws[coord].fill = styles.header_fill
        ws[coord].font = styles.header_font
        ws[coord].border = styles.border
    for idx, (left, right) in enumerate(rows, start=5):
        ws.cell(idx, 1, left).font = Font(bold=True)
        ws.cell(idx, 2, right).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(idx, 1).border = styles.border
        ws.cell(idx, 2).border = styles.border
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110
    ws.sheet_view.showGridLines = False


def _build_all(raw: pd.DataFrame, projection_source: object, market_source: object, top_n: int):
    """Roda toda a cadeia de transformacao e devolve as pecas + o resumo de KPIs."""
    weekly = build_weekly(raw)
    agg = build_agg(raw)
    agg, overall_conv, overall_ticket = add_classification(agg, raw)
    models = build_model_view(raw)
    projection, _, projection_target = load_projection(projection_source)
    market = load_market_csv(market_source)
    benchmark, multi_match = build_market_benchmark(raw, market)
    monthly_rate = float(projection["crescimento_mensal_necessario"].iloc[0])
    summary = build_summary(raw, agg, weekly, projection_target, monthly_rate, benchmark)
    notes = build_notes()
    action_plan = build_action_plan()

    kpis = {
        "periodo": f"{raw['period_start'].min():%d/%m/%Y} a {raw['period_end'].max():%d/%m/%Y}",
        "semanas": int(raw["file"].nunique()),
        "anuncios_unicos": int(raw["ad_id"].nunique()),
        "receita_total": float(raw["revenue"].sum()),
        "unidades": int(raw["units"].sum()),
        "pedidos": int(raw["orders"].sum()),
        "conversao_media": float(overall_conv),
        "ticket_medio": float(overall_ticket),
        "tem_benchmark": not benchmark.empty,
    }
    pieces = dict(
        weekly=weekly,
        agg=agg,
        models=models,
        projection=projection,
        summary=summary,
        notes=notes,
        action_plan=action_plan,
        benchmark=benchmark,
        multi_match=multi_match,
    )
    return pieces, kpis


def run_pipeline(
    report_files: Iterable[object],
    projection_file: object = None,
    market_file: object = None,
    top_n: int = 200,
) -> tuple[BytesIO, dict]:
    """Entrada do web app: recebe relatorios enviados (e arquivos opcionais), gera o Excel
    em memoria e devolve (buffer, kpis)."""
    raw = load_reports_from_files(report_files)
    pieces, kpis = _build_all(raw, projection_file, market_file, top_n)

    buffer = BytesIO()
    build_workbook(
        output=buffer,
        raw=raw,
        weekly=pieces["weekly"],
        agg=pieces["agg"],
        models=pieces["models"],
        projection=pieces["projection"],
        summary=pieces["summary"],
        notes=pieces["notes"],
        action_plan=pieces["action_plan"],
        benchmark=pieces["benchmark"],
        multi_match=pieces["multi_match"],
        top_n=top_n,
    )
    buffer.seek(0)
    return buffer, kpis


def main() -> None:
    args = parse_args()
    base_dir = Path.cwd()
    reports_dir = base_dir / args.reports_dir
    projection_file = base_dir / args.projection_file
    market_csv = base_dir / args.market_csv
    output_path = base_dir / args.output

    raw = load_reports(reports_dir)
    pieces, _ = _build_all(raw, projection_file, market_csv, args.top_n)

    build_workbook(
        output=output_path,
        raw=raw,
        weekly=pieces["weekly"],
        agg=pieces["agg"],
        models=pieces["models"],
        projection=pieces["projection"],
        summary=pieces["summary"],
        notes=pieces["notes"],
        action_plan=pieces["action_plan"],
        benchmark=pieces["benchmark"],
        multi_match=pieces["multi_match"],
        top_n=args.top_n,
    )
    print(output_path)


if __name__ == "__main__":
    main()
